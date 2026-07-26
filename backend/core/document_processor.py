"""
文档处理器模块
==============
处理文档上传后的完整流程：验证 → 加载 → 切分 → 向量化 → 存入ChromaDB。

支持的文档格式：
- PDF: 商品手册、规格说明书
- DOCX: Word文档
- TXT: 纯文本
- CSV: 表格数据（商品列表）
- XLSX: Excel表格（商品参数表）
- MD: Markdown文档
"""

import os
import uuid
from pathlib import Path
from typing import List, Optional
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.document_loaders import (
    PyPDFLoader,
    TextLoader,
    CSVLoader,
)
from langchain_core.documents import Document as LangchainDocument
from chromadb import PersistentClient
from chromadb.config import Settings as ChromaSettings
from backend.config import (
    UPLOAD_DIR,
    CHROMA_DIR,
    CHUNK_SIZE,
    CHUNK_OVERLAP,
    RETRIEVAL_TOP_K,
)
from backend.core.llm_factory import get_embeddings


# ChromaDB 持久化客户端（全局单例）
_chroma_client: Optional[PersistentClient] = None


def get_chroma_client() -> PersistentClient:
    """
    获取 ChromaDB 持久化客户端（单例模式）。

    什么是 ChromaDB？
    ChromaDB 是一个向量数据库。它专门存储"向量"（一串数字）和对应的原文。
    查询时，它能在毫秒级别找到和问题向量最接近的知识库片段。

    为什么要持久化？
    如果不持久化，数据只存在于内存中，服务重启后知识库全部丢失。
    持久化模式将数据存到磁盘，重启后数据还在。
    """
    global _chroma_client
    if _chroma_client is None:
        _chroma_client = PersistentClient(
            path=str(CHROMA_DIR),
            settings=ChromaSettings(anonymized_telemetry=False),
        )
    return _chroma_client


# LangChain 的 Chroma 向量存储封装
from langchain_community.vectorstores import Chroma


def get_or_create_collection(collection_name: str) -> Chroma:
    """
    获取或创建 ChromaDB 集合。

    每个知识库文档对应一个独立的 Collection（集合），
    这样做的好处：
    - 删除文档时可以精确删除对应的向量数据
    - 不同文档的向量数据互不干扰
    """
    embeddings = get_embeddings()
    client = get_chroma_client()

    collection = Chroma(
        client=client,
        collection_name=collection_name,
        embedding_function=embeddings,
    )
    return collection


def delete_collection(collection_name: str) -> None:
    """删除 ChromaDB 中的集合（用于文档删除时清理向量数据）"""
    try:
        client = get_chroma_client()
        client.delete_collection(collection_name)
    except Exception:
        pass  # 集合不存在时忽略


# ============================================================
# 文档加载
# ============================================================

def load_document(file_path: str) -> List[LangchainDocument]:
    """
    根据文件类型选择合适的 Loader 加载文档。

    LangChain 的 Loader 是什么？
    Loader 是"文档读取器"。不同格式的文件需要不同的读取方式：
    - PDF 需要 PyPDFLoader 来解析
    - Word 需要 Docx2txtLoader
    - 等等...
    LangChain 提供了统一的接口，屏蔽了底层的格式差异。
    """
    ext = Path(file_path).suffix.lower()

    if ext == ".pdf":
        loader = PyPDFLoader(file_path)
    elif ext == ".txt":
        loader = TextLoader(file_path, encoding="utf-8")
    elif ext == ".csv":
        loader = CSVLoader(file_path, encoding="utf-8")
    elif ext == ".md":
        # Markdown 本质是纯文本，直接用 TextLoader（避免依赖 unstructured 库）
        loader = TextLoader(file_path, encoding="utf-8")
    elif ext in [".docx", ".xlsx"]:
        # DOCX 和 XLSX 使用更底层的方式加载
        return _load_office_document(file_path, ext)
    else:
        raise ValueError(f"不支持的文件类型: {ext}")

    return loader.load()


def _load_office_document(file_path: str, ext: str) -> List[LangchainDocument]:
    """
    加载 Office 文档（docx, xlsx）。

    为什么单独处理？
    LangChain 对 Word/Excel 的支持不如 PDF 完善，需要手动处理。
    """
    content = ""

    if ext == ".docx":
        import docx2txt
        content = docx2txt.process(file_path)
    elif ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            content += f"\n=== Sheet: {sheet_name} ===\n"
            for row in ws.iter_rows(values_only=True):
                # 过滤全空行
                if any(cell is not None for cell in row):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    content += row_text + "\n"

    if not content.strip():
        raise ValueError(f"无法从文件中提取文本内容: {file_path}")

    filename = Path(file_path).name
    return [LangchainDocument(page_content=content, metadata={"source": filename})]


# ============================================================
# 文档切分
# ============================================================

def split_documents(documents: List[LangchainDocument]) -> List[LangchainDocument]:
    """
    将文档切分成小片段（Chunks）。

    为什么要切分？
    1. LLM 每次能处理的文本长度有限（Token 限制）
    2. 切分后每个小片段可以精确匹配问题，而非返回整个文档
    3. 小片段更容易做向量检索

    RecursiveCharacterTextSplitter 的工作方式：
    优先按段落(\\n\\n)切 → 然后按句子(\\n)切 → 最后按词切
    这样尽量保证每个片段是一个完整的语义单元，不会在句子中间截断。

    chunk_overlap（重叠）的作用：
    相邻片段之间有100字的重叠。避免一个问题正好跨越两个片段的边界，
    导致检索时丢失关键信息。
    """
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
        separators=["\n\n", "\n", "。", "！", "？", "，", " ", ""],
        length_function=len,
    )
    return text_splitter.split_documents(documents)


# ============================================================
# 文档入库流程
# ============================================================

def process_and_store_document(
    file_path: str,
    filename: str,
    collection_name: str,
    progress_callback=None,
) -> int:
    """
    处理文档的完整流程：加载 → 切分 → 向量化 → 存入ChromaDB。

    Args:
        file_path: 上传文件的完整路径
        filename: 原始文件名
        collection_name: ChromaDB 集合名称
        progress_callback: 可选，进度回调函数 (progress: float) -> None

    Returns:
        切分的片段数量
    """
    # 步骤1：加载文档
    if progress_callback:
        progress_callback(0.1)
    documents = load_document(file_path)

    # 步骤2：切分文档
    if progress_callback:
        progress_callback(0.3)
    chunks = split_documents(documents)

    if not chunks:
        raise ValueError("文档内容为空，无法切分")

    # 步骤3：为每个chunk补充元数据（来源信息）
    if progress_callback:
        progress_callback(0.5)
    for i, chunk in enumerate(chunks):
        chunk.metadata["filename"] = filename
        chunk.metadata["chunk_index"] = i
        # 如果有页码信息（PDF），保留
        if "page" not in chunk.metadata:
            chunk.metadata["page"] = chunk.metadata.get("page", 0)

    # 步骤4：向量化并存入 ChromaDB
    if progress_callback:
        progress_callback(0.7)
    embeddings = get_embeddings()
    client = get_chroma_client()

    from langchain_community.vectorstores import Chroma
    Chroma.from_documents(
        documents=chunks,
        embedding=embeddings,
        client=client,
        collection_name=collection_name,
    )

    if progress_callback:
        progress_callback(1.0)

    return len(chunks)


def get_collection_chunks(collection_name: str, limit: int = 50) -> List[dict]:
    """获取集合中的片段内容（用于预览）"""
    try:
        collection = get_or_create_collection(collection_name)
        results = collection.get(limit=limit)
        chunks = []
        if results and results["documents"]:
            for i, doc in enumerate(results["documents"]):
                meta = results["metadatas"][i] if results["metadatas"] else {}
                chunks.append({
                    "content": doc[:200],  # 只返回前200字预览
                    "full_content": doc,
                    "metadata": meta,
                })
        return chunks
    except Exception:
        return []


# 全局检索器（用于问答时跨文档检索）
_global_retriever: Optional[Chroma] = None


def get_global_collection() -> Chroma:
    """
    获取全局知识库集合（用于跨文档检索）。

    为什么需要一个全局集合？
    用户提问时，我们需要在所有文档中搜索，而不是指定某个文档。
    这个方法返回一个可以搜索所有知识库的 Chroma 实例。
    """
    global _global_retriever
    if _global_retriever is None:
        embeddings = get_embeddings()
        client = get_chroma_client()
        # 列出所有集合
        all_collections = client.list_collections()
        if all_collections:
            # 创建跨集合的检索
            _global_retriever = Chroma(
                client=client,
                collection_name=all_collections[0].name,
                embedding_function=embeddings,
            )
        else:
            # 如果没有集合，创建一个默认的
            _global_retriever = Chroma(
                client=client,
                collection_name="_default_",
                embedding_function=embeddings,
            )
    return _global_retriever
